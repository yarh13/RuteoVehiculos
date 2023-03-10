
import mysql.connector
from openpyxl import Workbook

# Coneccion a MYSQL
conexion1=mysql.connector.connect(host="localhost", user="root", passwd="", database="dbprueba")
cursor1=conexion1.cursor()
cursor1.execute("select * from crear_rutas")
cursoinfo= list(cursor1)
#for row in cursor1:
print(cursoinfo)

for i in range(13):
	(cursoinfo[2][i])
    
idRuta=print (cursoinfo[2][0])
idUserRuta=print (cursoinfo[2][1])
nombreRuta=print (cursoinfo[2][2])
canVehiRuta=print (cursoinfo[2][3])
horaInRuta=print (cursoinfo[2][4])
horaFinRuta=print (cursoinfo[2][5])
capCargRuta=print (cursoinfo[2][6])
canClientesRuta=print (cursoinfo[2][7])
horaInRecepRuta=print (cursoinfo[2][8])
horaFinRecepRuta=print (cursoinfo[2][9])
horaCreacionRuta=print (cursoinfo[2][10])

conexion1.close()


wb = Workbook()
from openpyxl import load_workbook 
wb = load_workbook(filename = 'ruteo_macro.xlsm')


# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['D192'] = idRuta
ws['D193'] = idUserRuta
ws['D194'] = nombreRuta
ws['D195'] = canVehiRuta
ws['D196'] = horaInRuta
ws['D197'] = horaFinRuta
ws['D198'] = capCargRuta
ws['D199'] = canClientesRuta
ws['D200'] = horaInRecepRuta
ws['D201'] = horaFinRecepRuta
ws['D201'] = horaCreacionRuta

# Save the file
wb.save("sample3.xlsm")